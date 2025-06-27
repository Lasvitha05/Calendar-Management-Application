# calendar_logic.py

import datetime
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from openpyxl.utils.exceptions import InvalidFileException 

class Event:
    def __init__(self, title, start_time, end_time):
        self.title = title
        self.start_time = start_time
        self.end_time = end_time

    def __repr__(self):
        return (f"Title: {self.title}, "
                f"Start: {self.start_time.strftime('%Y-%m-%d %H:%M')}, "
                f"End: {self.end_time.strftime('%Y-%m-%d %H:%M')}")

    def to_list(self):
        return [self.title, self.start_time, self.end_time]

class Calendar:
    def __init__(self, data_file='/Users/pregadalasvitha/Desktop/Calendar Application/Calendar StreamLit Code/calendar_data.xlsx'):
        self.events = []
        self.data_file = data_file
        self.sheet_name = 'Events'
        self._load_events()

    def _load_events(self):
        self.events.clear()
        try:
            if os.path.exists(self.data_file):
                wb = load_workbook(self.data_file)
                sheet = wb[self.sheet_name] if self.sheet_name in wb.sheetnames else wb.active
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if all(cell is None for cell in row): continue
                    try:
                        title, start_time, end_time = row
                        if isinstance(start_time, datetime.datetime) and isinstance(end_time, datetime.datetime):
                            self.events.append(Event(title, start_time, end_time))
                    except: continue
                self.events.sort(key=lambda e: e.start_time)
                self.workbook = wb
                self.sheet = sheet
            else:
                self.workbook = Workbook()
                self.sheet = self.workbook.active
                self.sheet.title = self.sheet_name
                self.sheet.append(["Title", "Start Time", "End Time"])
                for col_idx, cell in enumerate(self.sheet[1]):
                    cell.font = Font(bold=True)
                    self.sheet.column_dimensions[get_column_letter(col_idx+1)].width = 25
                self._save_workbook()
        except:
            self._recreate_empty_workbook()

    def _recreate_empty_workbook(self):
        self.events = []
        self.workbook = Workbook()
        self.sheet = self.workbook.active
        self.sheet.title = self.sheet_name
        self.sheet.append(["Title", "Start Time", "End Time"])
        self._save_workbook()

    def _save_workbook(self):
        try:
            self.workbook.save(self.data_file)
        except:
            pass

    def create_event(self, title, start_time_str, end_time_str):
        try:
            start_time = datetime.datetime.strptime(start_time_str, '%Y-%m-%d %H:%M')
            end_time = datetime.datetime.strptime(end_time_str, '%Y-%m-%d %H:%M')
        except ValueError:
            return False, "Invalid datetime format."

        if start_time >= end_time:
            return False, "Start time must be before end time."

        for e in self.events:
            if (start_time < e.end_time and end_time > e.start_time):
                return False, f"Overlaps with: {e.title}"

        new_event = Event(title, start_time, end_time)
        self.events.append(new_event)
        self.events.sort(key=lambda e: e.start_time)
        self.sheet.append(new_event.to_list())
        self._save_workbook()
        return True, "Event created."

    def delete_event(self, index):
        try:
            index = int(index) - 1
            if not (0 <= index < len(self.events)): return False
            del self.events[index]
            self.sheet.delete_rows(2, self.sheet.max_row)
            for e in self.events:
                self.sheet.append(e.to_list())
            self._save_workbook()
            return True
        except:
            return False

    def list_events_for_day(self, date_str):
        try:
            target_date = datetime.datetime.strptime(date_str, '%Y-%m-%d').date()
            return [e for e in self.events if e.start_time.date() == target_date]
        except:
            return []

    def list_remaining_events_for_today(self):
        now = datetime.datetime.now()
        return [e for e in self.events if e.start_time.date() == now.date() and e.end_time > now]

    def find_all_available_slots(self, duration_minutes, target_date):
        duration = datetime.timedelta(minutes=duration_minutes)
        if target_date == datetime.datetime.now().date():
            search_start = datetime.datetime.now()
        else:
            search_start = datetime.datetime.combine(target_date, datetime.time(0, 0))

        search_end = datetime.datetime.combine(target_date, datetime.time(23, 59, 59))
        slots = []
        current = search_start
        events = sorted([e for e in self.events if e.start_time.date() == target_date], key=lambda e: e.start_time)

        for e in events:
            gap_start = current
            gap_end = e.start_time
            while gap_start + duration <= gap_end:
                slots.append((gap_start, gap_start + duration))
                gap_start += datetime.timedelta(minutes=1)
            current = max(current, e.end_time)

        while current + duration <= search_end:
            slots.append((current, current + duration))
            current += datetime.timedelta(minutes=1)

        return slots
